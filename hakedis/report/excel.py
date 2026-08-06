"""Excel metraj cetveli (openpyxl).

Uretilen sayfalar:
  Ozet          - tip bazinda beton/kalip toplamlari, hesap parametreleri
  Metraj Cetveli- klasik kirik olcu cetveli (poz/en/boy/yukseklik/alan/hacim)
  Kirik Olcu    - her elemanin olcu kirilimi, koordinatlariyla
  Elemanlar     - tespit edilen elemanlarin ham dokumu
  Uyarilar      - kontrol edilmesi gereken noktalar

Cetveldeki HER hucre formul olarak degil, hesaplanmis deger olarak yazilir;
ancak "Formul" sutunu olcunun nasil alindigini metin olarak tasir, boylece
kontrol muhendisi her satiri elle dogrulayabilir.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from hakedis.model import ElemanTipi, MetrajSonucu

BASLIK_DOLGU = PatternFill("solid", fgColor="1F4E79")
BASLIK_YAZI = Font(bold=True, color="FFFFFF", size=10)
GRUP_DOLGU = PatternFill("solid", fgColor="DDEBF7")
DUSUM_YAZI = Font(color="C00000", italic=True, size=10)
DUSUM_DOLGU = PatternFill("solid", fgColor="FCE4E4")
TOPLAM_YAZI = Font(bold=True, size=10)
TOPLAM_DOLGU = PatternFill("solid", fgColor="FFF2CC")
UYARI_YAZI = Font(color="C00000", size=10)

INCE = Side(style="thin", color="B0B0B0")
CERCEVE = Border(left=INCE, right=INCE, top=INCE, bottom=INCE)

SUTUNLAR = [
    ("Poz No", 16),
    ("Eleman", 12),
    ("Tanim", 34),
    ("Benzer", 8),
    ("En (m)", 10),
    ("Boy (m)", 10),
    ("Yuks. (m)", 10),
    ("Alan (m2)", 12),
    ("Hacim (m3)", 12),
    ("Demir (kg)", 12),
    ("Birim", 7),
    ("Olcu Aciklamasi (kirik olcu)", 52),
]


def _baslik_yaz(ws, sutunlar, satir: int = 1) -> int:
    for i, (ad, genislik) in enumerate(sutunlar, start=1):
        h = ws.cell(row=satir, column=i, value=ad)
        h.fill = BASLIK_DOLGU
        h.font = BASLIK_YAZI
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        h.border = CERCEVE
        ws.column_dimensions[get_column_letter(i)].width = genislik
    ws.freeze_panes = ws.cell(row=satir + 1, column=1)
    return satir + 1


def _sayfa_ozet(wb: Workbook, sonuc: MetrajSonucu) -> None:
    ws = wb.create_sheet("Ozet", 0)
    ws.column_dimensions["A"].width = 26
    for h in "BCDEF":
        ws.column_dimensions[h].width = 14

    ws["A1"] = "KIRIK OLCU METRAJ OZETI"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Kaynak dosya: {Path(sonuc.kaynak_dosya).name}"
    ws["A3"] = f"Kat: {sonuc.kat}"

    satir = 5
    ws.cell(row=satir, column=1, value="Hesap parametreleri").font = TOPLAM_YAZI
    satir += 1
    etiketler = {
        "kat_yuksekligi": "Kat yuksekligi (m)",
        "doseme_kalinligi": "Doseme kalinligi (m)",
        "net_yukseklik": "Kolon/perde net yuksekligi (m)",
        "birim": "Cizim birimi",
    }
    for anahtar, etiket in etiketler.items():
        if anahtar in sonuc.parametreler:
            ws.cell(row=satir, column=1, value=etiket)
            ws.cell(row=satir, column=2, value=sonuc.parametreler[anahtar])
            satir += 1

    satir += 1
    basliklar = ["Eleman Tipi", "Adet", "Beton (m3)", "Kalip (m2)", "Demir (kg)"]
    for i, b in enumerate(basliklar, start=1):
        h = ws.cell(row=satir, column=i, value=b)
        h.fill = BASLIK_DOLGU
        h.font = BASLIK_YAZI
        h.border = CERCEVE
    satir += 1

    ozet = sonuc.ozet()
    beton_toplam = kalip_toplam = demir_toplam = 0.0
    for tip in ElemanTipi:
        k = ozet.get(tip.value)
        if not k or (
            k["adet"] == 0 and k["beton_m3"] == 0 and k["kalip_m2"] == 0
        ):
            continue
        ws.cell(row=satir, column=1, value=tip.value).border = CERCEVE
        ws.cell(row=satir, column=2, value=int(k["adet"])).border = CERCEVE
        ws.cell(row=satir, column=3, value=round(k["beton_m3"], 3)).border = CERCEVE
        ws.cell(row=satir, column=4, value=round(k["kalip_m2"], 3)).border = CERCEVE
        ws.cell(row=satir, column=5, value=round(k["demir_kg"], 2)).border = CERCEVE
        beton_toplam += k["beton_m3"]
        kalip_toplam += k["kalip_m2"]
        demir_toplam += k["demir_kg"]
        satir += 1

    for i, deger in enumerate(
        ["TOPLAM", "", round(beton_toplam, 3), round(kalip_toplam, 3), round(demir_toplam, 2)],
        start=1,
    ):
        h = ws.cell(row=satir, column=i, value=deger)
        h.font = TOPLAM_YAZI
        h.fill = TOPLAM_DOLGU
        h.border = CERCEVE

    satir += 3
    ws.cell(
        row=satir,
        column=1,
        value=(
            "Bu cetvel otomatik uretilmistir. Teslim etmeden once SVG kontrol "
            "paftasini inceleyip 'Uyarilar' sayfasini okuyun."
        ),
    ).font = UYARI_YAZI


def _sayfa_cetvel(wb: Workbook, sonuc: MetrajSonucu) -> None:
    ws = wb.create_sheet("Metraj Cetveli")
    satir = _baslik_yaz(ws, SUTUNLAR)

    sira = [
        ElemanTipi.KOLON,
        ElemanTipi.PERDE,
        ElemanTipi.KIRIS,
        ElemanTipi.DOSEME,
        ElemanTipi.MERDIVEN,
        ElemanTipi.DUVAR,
        ElemanTipi.KAPI,
        ElemanTipi.PENCERE,
        ElemanTipi.BILINMEYEN,
    ]
    grup_baslik = {
        ElemanTipi.DUVAR: "DUVAR (BOLME) METRAJI",
        ElemanTipi.KAPI: "KAPI-PENCERE METRAJI (DOGRAMA LISTESI)",
        ElemanTipi.PENCERE: "KAPI-PENCERE METRAJI (DOGRAMA LISTESI)",
        ElemanTipi.BILINMEYEN: "SIVA / KAPLAMA (BETONARME YUZEYLERI)",
    }
    # Kat gruplari: cok katli birlesik sonucta kat, sonra tip sirasiyla goster
    kats = list(dict.fromkeys(s.kat for s in sonuc.satirlar if s.kat)) or [""]
    for kat in kats:
        kat_satirlari = [s for s in sonuc.satirlar if s.kat == kat]
        if len(kats) > 1:
            h = ws.cell(row=satir, column=1, value=f"KAT: {kat}")
            h.font = Font(bold=True, size=11)
            h.fill = BASLIK_DOLGU
            for i in range(1, len(SUTUNLAR) + 1):
                ws.cell(row=satir, column=i).fill = BASLIK_DOLGU
            satir += 1
        for tip in sira:
            satirlar = [s for s in kat_satirlari if s.tip == tip]
            if not satirlar:
                continue
            h = ws.cell(
                row=satir,
                column=1,
                value=grup_baslik.get(tip, f"{tip.value.upper()} METRAJI"),
            )
            h.font = Font(bold=True, size=11)
            h.fill = GRUP_DOLGU
            for i in range(1, len(SUTUNLAR) + 1):
                ws.cell(row=satir, column=i).fill = GRUP_DOLGU
            satir += 1

            beton = kalip = demir = 0.0
            for s in satirlar:
                isaret = -1.0 if s.dusum_mu else 1.0
                degerler = [
                    s.poz,
                    s.eleman_adi,
                    s.tanim,
                    s.benzer,
                    s.en,
                    s.boy,
                    s.yukseklik,
                    (isaret * s.alan) if s.alan is not None else None,
                    (isaret * s.hacim) if s.hacim is not None else None,
                    (isaret * s.kg) if s.kg is not None else None,
                    s.birim,
                    s.formul,
                ]
                for i, deger in enumerate(degerler, start=1):
                    h = ws.cell(row=satir, column=i, value=deger)
                    h.border = CERCEVE
                    if s.dusum_mu:
                        h.font = DUSUM_YAZI
                    if i in (5, 6, 7, 8, 9):
                        h.number_format = "0.000"
                    if i == 10:
                        h.number_format = "0.0"
                if s.birim == "m3":
                    beton += isaret * (s.hacim or 0.0)
                elif s.birim == "m2":
                    kalip += isaret * (s.alan or 0.0)
                elif s.birim == "kg":
                    demir += isaret * (s.kg or 0.0)
                satir += 1

            ws.cell(row=satir, column=3, value=f"{tip.value} ARA TOPLAM").font = TOPLAM_YAZI
            ws.cell(row=satir, column=8, value=round(kalip, 3)).font = TOPLAM_YAZI
            ws.cell(row=satir, column=9, value=round(beton, 3)).font = TOPLAM_YAZI
            ws.cell(row=satir, column=10, value=round(demir, 2)).font = TOPLAM_YAZI
            for i in range(1, len(SUTUNLAR) + 1):
                h = ws.cell(row=satir, column=i)
                h.fill = TOPLAM_DOLGU
                h.border = CERCEVE
                if i in (8, 9):
                    h.number_format = "0.000"
                if i == 10:
                    h.number_format = "0.0"
            satir += 2


def _sayfa_kirik_olcu(wb: Workbook, sonuc: MetrajSonucu) -> None:
    """Her olcunun nereden geldigini gosteren kirilim sayfasi."""
    ws = wb.create_sheet("Kirik Olcu")
    sutunlar = [
        ("Eleman", 12),
        ("Tip", 11),
        ("Kalem", 30),
        ("Olcu kirilimi", 90),
        ("Sonuc", 14),
        ("Birim", 7),
    ]
    satir = _baslik_yaz(ws, sutunlar)

    for s in sonuc.satirlar:
        if not s.detay:
            continue
        bas = satir
        ws.cell(row=satir, column=1, value=s.eleman_adi).font = TOPLAM_YAZI
        ws.cell(row=satir, column=2, value=s.tip.value)
        ws.cell(row=satir, column=3, value=s.tanim.strip())
        h = ws.cell(row=satir, column=5, value=round(s.miktar, 4))
        h.number_format = "0.0000"
        h.font = TOPLAM_YAZI
        ws.cell(row=satir, column=6, value=s.birim)
        for metin in s.detay:
            ws.cell(row=satir, column=4, value=metin).alignment = Alignment(
                horizontal="left"
            )
            satir += 1
        satir = max(satir, bas + 1)
        for r in range(bas, satir):
            for c in range(1, len(sutunlar) + 1):
                ws.cell(row=r, column=c).border = CERCEVE
        satir += 1


def _sayfa_elemanlar(wb: Workbook, sonuc: MetrajSonucu) -> None:
    ws = wb.create_sheet("Elemanlar")
    sutunlar = [
        ("Ad", 12),
        ("Tip", 11),
        ("Kat", 14),
        ("Kaynak katman", 18),
        ("Etiket metni", 22),
        ("b (m)", 9),
        ("h / t (m)", 10),
        ("Uzunluk (m)", 12),
        ("Alan (m2)", 12),
        ("Guven", 8),
        ("Notlar", 60),
    ]
    satir = _baslik_yaz(ws, sutunlar)
    for e in sonuc.elemanlar:
        degerler = [
            e.ad,
            e.tip.value,
            e.kat,
            e.kaynak_katman,
            e.etiket_metni,
            e.olculer.get("b"),
            e.olculer.get("h", e.olculer.get("t")),
            round(e.toplam_uzunluk, 3) if e.segmentler else None,
            e.olculer.get("net_alan", e.olculer.get("kesit_alani")),
            round(e.guven, 2),
            " | ".join(e.notlar),
        ]
        for i, deger in enumerate(degerler, start=1):
            h = ws.cell(row=satir, column=i, value=deger)
            h.border = CERCEVE
            if i in (6, 7, 8, 9):
                h.number_format = "0.000"
            if e.guven < 0.7:
                h.font = UYARI_YAZI
        satir += 1


def _sayfa_uyarilar(wb: Workbook, sonuc: MetrajSonucu) -> None:
    ws = wb.create_sheet("Uyarilar")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 130
    ws["A1"] = "#"
    ws["B1"] = "Kontrol edilmesi gereken noktalar"
    for h in (ws["A1"], ws["B1"]):
        h.fill = BASLIK_DOLGU
        h.font = BASLIK_YAZI
    if not sonuc.uyarilar:
        ws["B2"] = "Uyari yok. Yine de SVG kontrol paftasini inceleyin."
        return
    for i, u in enumerate(sonuc.uyarilar, start=1):
        ws.cell(row=i + 1, column=1, value=i)
        h = ws.cell(row=i + 1, column=2, value=u)
        h.font = UYARI_YAZI
        h.alignment = Alignment(wrap_text=True, vertical="top")


def _sayfa_maliyet_ozet(wb: Workbook, m: dict | None) -> None:
    """YIGS yaklasik maliyet ozet tutanagi (tek sayfa, ihale dokumani)."""
    ws = wb.create_sheet("YIGS Ozet")
    for harf, g in {"A": 34, "B": 14, "C": 16, "D": 60}.items():
        ws.column_dimensions[harf].width = g
    ws["A1"] = "YAKLASIK MALIYET OZET TUTANAGI"
    ws["A1"].font = Font(bold=True, size=14)

    if m is None:
        ws["A3"] = "Yaklasik maliyet devre disi."
        return

    ws.cell(row=3, column=1, value=f"Betonarme sinifi: {m.get('beton_sinifi', '?')}").font = Font(bold=True)
    ws.cell(row=4, column=1, value=f"Para birimi: {m['para_birimi']}").font = Font(bold=True)
    satir = 6
    basliklar = [("Bolum", "A"), ("Kalem", "B"), ("Tutar", "C"), ("Aciklama", "D")]
    for i, (ad, _h) in enumerate(basliklar, start=1):
        h = ws.cell(row=satir, column=i, value=ad)
        h.fill = BASLIK_DOLGU
        h.font = BASLIK_YAZI
        h.border = CERCEVE
    satir += 1

    bolumler: dict[str, list] = {}
    for k in m["kalemler"]:
        bolumler.setdefault(k["bolum"], []).append(k)

    for bolum, kalemler in bolumler.items():
        tutar = sum(k["tutar"] for k in kalemler)
        ws.cell(row=satir, column=1, value=bolum).font = Font(bold=True, size=11)
        ws.cell(row=satir, column=2, value=len(kalemler)).border = CERCEVE
        ws.cell(row=satir, column=3, value=round(tutar, 2)).border = CERCEVE
        ws.cell(row=satir, column=3).number_format = "0.00"
        ws.cell(row=satir, column=4).value = "; ".join(
            f"{k['poz']} {k['miktar']:g} {k['birim']}" for k in kalemler[:4]
        ) + (" …" if len(kalemler) > 4 else "")
        for c in range(1, 5):
            ws.cell(row=satir, column=c).border = CERCEVE
            ws.cell(row=satir, column=c).fill = GRUP_DOLGU
        satir += 1

    satir += 1
    for ad, deger in [
        ("ARA TOPLAM", m["ara_toplam"]),
        (f"KDV (%{m['kdv_oran']:g})", m["kdv"]),
        ("GENEL TOPLAM", m["genel_toplam"]),
    ]:
        ws.cell(row=satir, column=1, value=ad).font = TOPLAM_YAZI
        h = ws.cell(row=satir, column=3, value=round(deger, 2))
        h.font = TOPLAM_YAZI
        h.number_format = "0.00"
        for c in range(1, 5):
            ws.cell(row=satir, column=c).border = CERCEVE
        satir += 1

    if m["fiyatsiz_pozlar"]:
        satir += 1
        ws.cell(
            row=satir, column=1,
            value="Fiyat tanimsiz pozlar (hesaba dahil degil): "
            + ", ".join(m["fiyatsiz_pozlar"]),
        ).font = UYARI_YAZI
    satir += 1
    ws.cell(row=satir, column=1, value=m["not"]).alignment = Alignment(
        wrap_text=True, vertical="top"
    )


def _sayfa_maliyet(
    wb: Workbook, sonuc: MetrajSonucu, m: dict | None
) -> None:
    """Maliyet tablosunu ayri bir Excel sayfasi olarak yazar."""
    ws = wb.create_sheet("Maliyet")
    genislikler = {
        "A": 6, "B": 14, "C": 44, "D": 8, "E": 12, "F": 12, "G": 16,
    }
    for harf, g in genislikler.items():
        ws.column_dimensions[harf].width = g

    if m is None:
        ws["A1"] = "Yaklasik maliyet devre disi."
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = (
            "yapilandirmada maliyet.aktif: true yapin veya "
            "web arayuzunde 'Maliyet' sekmesini kullanin."
        )
        return

    ws["A1"] = f"YAKLASIK MALIYET ({m['para_birimi']}) - YIGS FORMATI"
    ws["A1"].font = Font(bold=True, size=14)
    satir = 3
    basliklar = [
        ("Sira No", "A"), ("Poz No", "B"), ("Poz Tanimi", "C"),
        ("Birim", "D"), ("Miktar", "E"), ("Birim Fiyat", "F"),
        ("Tutar", "G"),
    ]
    for i, (ad, _harf) in enumerate(basliklar, start=1):
        h = ws.cell(row=satir, column=i, value=ad)
        h.fill = BASLIK_DOLGU
        h.font = BASLIK_YAZI
        h.border = CERCEVE
    satir += 1

    son_bolum = None
    bolum_toplam: dict[str, float] = {}
    for k in m["kalemler"]:
        if k["bolum"] != son_bolum:
            if son_bolum is not None:
                satir += 1
                ws.cell(
                    row=satir, column=2, value=f"{son_bolum} ARA TOPLAM"
                ).font = TOPLAM_YAZI
                ws.cell(
                    row=satir, column=7, value=round(bolum_toplam.get(son_bolum, 0.0), 2)
                ).font = TOPLAM_YAZI
                ws.cell(row=satir, column=7).number_format = "0.00"
                for c in range(1, 8):
                    ws.cell(row=satir, column=c).fill = GRUP_DOLGU
            satir += 1
            son_bolum = k["bolum"]
            bolum_toplam.setdefault(son_bolum, 0.0)
            ws.cell(row=satir, column=1, value=son_bolum).font = Font(bold=True, size=11)
            ws.cell(row=satir, column=1).fill = GRUP_DOLGU
            for c in range(1, 8):
                ws.cell(row=satir, column=c).fill = GRUP_DOLGU
        bolum_toplam[son_bolum] = bolum_toplam.get(son_bolum, 0.0) + k["tutar"]
        degerler = [
            k["sira"],
            k["poz"],
            k["tanim"],
            k["birim"],
            k["miktar"],
            k["fiyat"],
            k["tutar"],
        ]
        for i, deger in enumerate(degerler, start=1):
            h = ws.cell(row=satir, column=i, value=deger)
            h.border = CERCEVE
            if i in (5, 6, 7):
                h.number_format = "0.00"
        if k["dusum"]:
            for i in range(1, 8):
                ws.cell(row=satir, column=i).fill = DUSUM_DOLGU
        satir += 1

    if son_bolum is not None:
        satir += 1
        ws.cell(row=satir, column=2, value=f"{son_bolum} ARA TOPLAM").font = TOPLAM_YAZI
        ws.cell(
            row=satir, column=7, value=round(bolum_toplam.get(son_bolum, 0.0), 2)
        ).font = TOPLAM_YAZI
        ws.cell(row=satir, column=7).number_format = "0.00"
        for c in range(1, 8):
            ws.cell(row=satir, column=c).fill = GRUP_DOLGU

    satir += 1
    for ad, deger in [
        ("ARA TOPLAM", m["ara_toplam"]),
        (f"KDV (%{m['kdv_oran']:g})", m["kdv"]),
        ("GENEL TOPLAM", m["genel_toplam"]),
    ]:
        ws.cell(row=satir, column=2, value=ad).font = TOPLAM_YAZI
        ws.cell(row=satir, column=7, value=round(deger, 2)).font = TOPLAM_YAZI
        ws.cell(row=satir, column=7).number_format = "0.00"
        for c in range(1, 8):
            ws.cell(row=satir, column=c).border = CERCEVE
        satir += 1

    if m["fiyatsiz_pozlar"]:
        satir += 1
        h = ws.cell(
            row=satir, column=1,
            value="Fiyat tanimsiz pozlar: " + ", ".join(m["fiyatsiz_pozlar"]),
        )
        h.font = UYARI_YAZI
    satir += 1
    ws.cell(row=satir, column=1, value=m["not"]).alignment = Alignment(
        wrap_text=True, vertical="top"
    )


def excel_yaz(
    sonuc: MetrajSonucu,
    hedef: str | Path,
    ayarlar=None,
) -> Path:
    """Metraj sonucunu Excel dosyasi olarak yazar."""
    hedef = Path(hedef)
    wb = Workbook()
    wb.remove(wb.active)

    _sayfa_ozet(wb, sonuc)
    _sayfa_cetvel(wb, sonuc)
    _sayfa_kirik_olcu(wb, sonuc)
    _sayfa_elemanlar(wb, sonuc)
    _sayfa_uyarilar(wb, sonuc)
    if ayarlar is not None:
        from hakedis.maliyet import maliyet_hesapla

        if ayarlar.al("maliyet.aktif", False):
            m = maliyet_hesapla(sonuc, ayarlar)
            _sayfa_maliyet(wb, sonuc, m)
            _sayfa_maliyet_ozet(wb, m)
        else:
            _sayfa_maliyet(wb, sonuc, None)
            _sayfa_maliyet_ozet(wb, None)

    hedef.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(hedef))
    return hedef


def _topla(satirlar) -> tuple[float, float, float]:
    """Bir satir grubunun beton/kalip/demir toplamini dondurur."""
    beton = kalip = demir = 0.0
    for s in satirlar:
        isaret = -1.0 if s.dusum_mu else 1.0
        if s.birim == "m3":
            beton += isaret * (s.hacim or 0.0)
        elif s.birim == "m2":
            kalip += isaret * (s.alan or 0.0)
        elif s.birim == "kg":
            demir += isaret * (s.kg or 0.0)
    return beton, kalip, demir


def _sayfa_kat_ozeti(wb: Workbook, sonuc: MetrajSonucu) -> None:
    """Cok katli birlesik sonucta kat x eleman tipi ozetini yazar."""
    ws = wb.create_sheet("Kat Ozeti", 0)
    genislikler = {"A": 24, "B": 14, "C": 8, "D": 14, "E": 14, "F": 14}
    for harf, g in genislikler.items():
        ws.column_dimensions[harf].width = g

    ws["A1"] = "KAT BAZINDA KIRIK OLCU METRAJ OZETI"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Kaynak: {Path(sonuc.kaynak_dosya).name if sonuc.kaynak_dosya else ''}"

    satir = 4
    for i, b in enumerate(["Kat", "Eleman Tipi", "Adet", "Beton (m3)", "Kalip (m2)", "Demir (kg)"], start=1):
        h = ws.cell(row=satir, column=i, value=b)
        h.fill = BASLIK_DOLGU
        h.font = BASLIK_YAZI
        h.border = CERCEVE
    satir += 1

    kats = list(dict.fromkeys(s.kat for s in sonuc.satirlar if s.kat))
    sira = [
        ElemanTipi.KOLON,
        ElemanTipi.PERDE,
        ElemanTipi.KIRIS,
        ElemanTipi.DOSEME,
        ElemanTipi.MERDIVEN,
    ]
    if not kats:
        return
    for kat in kats:
        kat_satirlari = [s for s in sonuc.satirlar if s.kat == kat]
        kat_elemanlar = [e for e in sonuc.elemanlar if e.kat == kat]
        for tip in sira:
            grubu = [s for s in kat_satirlari if s.tip == tip]
            if not grubu:
                continue
            beton, kalip, demir = _topla(grubu)
            adet = sum(1 for e in kat_elemanlar if e.tip == tip)
            degerler = [kat, tip.value, adet, round(beton, 3), round(kalip, 3), round(demir, 2)]
            for i, deger in enumerate(degerler, start=1):
                h = ws.cell(row=satir, column=i, value=deger)
                h.border = CERCEVE
                if i in (4, 5):
                    h.number_format = "0.000"
                if i == 6:
                    h.number_format = "0.0"
            satir += 1
        bt, kt, dt = _topla(kat_satirlari)
        for i, deger in enumerate([kat, "KAT TOPLAM", "", round(bt, 3), round(kt, 3), round(dt, 2)], start=1):
            h = ws.cell(row=satir, column=i, value=deger)
            h.font = TOPLAM_YAZI
            h.fill = TOPLAM_DOLGU
            h.border = CERCEVE
            if i in (4, 5):
                h.number_format = "0.000"
            if i == 6:
                h.number_format = "0.0"
        satir += 1

    bt, kt, dt = _topla(sonuc.satirlar)
    for i, deger in enumerate(["TOPLAM", "", "", round(bt, 3), round(kt, 3), round(dt, 2)], start=1):
        h = ws.cell(row=satir, column=i, value=deger)
        h.font = Font(bold=True, size=12)
        h.fill = GRUP_DOLGU
        h.border = CERCEVE
        if i in (4, 5):
            h.number_format = "0.000"
        if i == 6:
            h.number_format = "0.0"


def _sayfa_maliyet_kat(
    wb: Workbook, sonuclar: list[MetrajSonucu], ayarlar
) -> None:
    """Kat bazli yaklasik maliyet karsilastirma sayfasi."""
    from hakedis.maliyet import maliyet_hesapla

    ws = wb.create_sheet("Maliyet Kat")
    for harf, g in {"A": 22, "B": 14, "C": 14, "D": 16, "E": 34}.items():
        ws.column_dimensions[harf].width = g
    ws["A1"] = "KAT BAZINDA YAKLASIK MALIYET"
    ws["A1"].font = Font(bold=True, size=14)
    satir = 3
    basliklar = [
        ("Kat", "A"), ("Ara Toplam", "B"), ("KDV", "C"),
        ("Genel Toplam", "D"), ("Not", "E"),
    ]
    for i, (ad, _h) in enumerate(basliklar, start=1):
        h = ws.cell(row=satir, column=i, value=ad)
        h.fill = BASLIK_DOLGU
        h.font = BASLIK_YAZI
        h.border = CERCEVE
    satir += 1

    toplamlar = [0.0, 0.0, 0.0]
    for s in sonuclar:
        m = maliyet_hesapla(s, ayarlar)
        eksik = ", ".join(m["fiyatsiz_pozlar"]) if m["fiyatsiz_pozlar"] else ""
        degerler = [
            s.kat or "?",
            m["ara_toplam"],
            m["kdv"],
            m["genel_toplam"],
            eksik,
        ]
        for i, deger in enumerate(degerler, start=1):
            h = ws.cell(row=satir, column=i, value=deger)
            h.border = CERCEVE
            if i in (2, 3, 4):
                h.number_format = "0.00"
        if eksik:
            ws.cell(row=satir, column=5).font = UYARI_YAZI
        for i, t in enumerate((m["ara_toplam"], m["kdv"], m["genel_toplam"]), start=1):
            toplamlar[i - 1] += t
        satir += 1

    degerler = ["TOPLAM"] + [round(t, 2) for t in toplamlar] + [""]
    for i, deger in enumerate(degerler, start=1):
        h = ws.cell(row=satir, column=i, value=deger)
        h.font = Font(bold=True, size=12)
        h.fill = GRUP_DOLGU
        h.border = CERCEVE
        if i in (2, 3, 4):
            h.number_format = "0.00"


def excel_yaz_toplu(
    sonuclar: list[MetrajSonucu], hedef: str | Path, ayarlar=None
) -> Path:
    """Cok katli/paftali calismanin ortak Excel dosyasini yazar.

    Kat Ozeti ilk sayfada; metraj cetveli/eleman/uyari sayfalari kat bazli
    kirilimlarla gelir. `ayarlar` verilir ve `maliyet.aktif` ise YIGS
    formatinda yaklasik maliyet + kat bazli karsilastirma sayfalari eklenir.
    Bireysel katlar icin `--kat-adlari` (cok dosya) veya `--paftalar`
    (cok PDF sayfasi) kullanilir.
    """
    from hakedis.maliyet import maliyet_hesapla
    from hakedis.metraj import sonuclari_birlestir

    hedef = Path(hedef)
    birlesik = sonuclari_birlestir(sonuclar)
    wb = Workbook()
    wb.remove(wb.active)

    _sayfa_kat_ozeti(wb, birlesik)
    _sayfa_cetvel(wb, birlesik)
    _sayfa_kirik_olcu(wb, birlesik)
    _sayfa_elemanlar(wb, birlesik)
    _sayfa_uyarilar(wb, birlesik)
    if ayarlar is not None:
        if ayarlar.al("maliyet.aktif", False):
            m = maliyet_hesapla(birlesik, ayarlar)
            _sayfa_maliyet(wb, birlesik, m)
            _sayfa_maliyet_kat(wb, sonuclar, ayarlar)
            _sayfa_maliyet_ozet(wb, m)
        else:
            _sayfa_maliyet(wb, birlesik, None)
            _sayfa_maliyet_ozet(wb, None)

    hedef.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(hedef))
    return hedef
