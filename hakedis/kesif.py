"""Alinan kesif (poz gruplu metraj + birim fiyat) raporu.

`maliyet.py` metraj satir basi kalem uretir (her eleman ayri); kesif ise
Ayni poz ve imalat tanimindaki miktarlari tek satirda toplayarak ihale
teslim belgesi duzeninde hazirlar:

  No | Poz No | Imalat (poz tanimi) | Birim | Miktar | Birim Fiyat | Tutar

Kalemler bolumlere (betonarme, siva, dograma, kaplama) ayrilir, dusum
satirlari eksili yazilir ve ara/Kdv/genel toplam cikarilir. Fiyatsiz pozlar
miktarlariyla birlikte listelenir (fiyat ve tutar bos) ve uyarida bildirilir.

Standalone `kesif_excel_yaz` disinda, ana metraj workbook'unu da besleyen
`kesif_sayfasi_yaz` ayni cetveli var olan `wb` içine ekler.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from hakedis.config import Ayarlar
from hakedis.maliyet import BOLUMLER, _bolum, fiyat_sozlugu
from hakedis.model import MetrajSonucu

BASLIK_DOLGU = PatternFill("solid", fgColor="1F4E79")
BASLIK_YAZI = Font(bold=True, color="FFFFFF", size=10)
GRUP_DOLGU = PatternFill("solid", fgColor="DDEBF7")
DUSUM_YAZI = Font(color="C00000", italic=True, size=10)
DUSUM_DOLGU = PatternFill("solid", fgColor="FCE4E4")
TOPLAM_YAZI = Font(bold=True, size=10)
TOPLAM_DOLGU = PatternFill("solid", fgColor="FFF2CC")
UYARI_YAZI = Font(color="C00000", size=10)

INCE = Side(style="thin", color="B0B0B0")
CERCEVE = Border(left=INCE, right=INCE, top=INCE, bottom=INCE)


def kesif_hesapla(sonuc: MetrajSonucu, ayarlar: Ayarlar) -> dict:
    """Metraj satirlarini poz bazinda toplayip alinan kesif tablosu uretir.

    Ayni (poz, imalat tanimi, birim) grubundaki miktarlar toplanir. Dusum
    satirlari isaretle toplanir. Fiyatsiz pozlar da miktarlariyla listelenir;
    tutar/fiyat None kalir.
    """
    fiyatlar = fiyat_sozlugu(ayarlar)
    kdv_oran = float(ayarlar.al("maliyet.kdv_oran", 20))
    sinif = str(ayarlar.al("kat.beton_sinifi", "C25/30") or "C25/30")

    gruplar: dict[tuple, dict] = {}
    for s in sonuc.satirlar:
        miktar = s.miktar
        if miktar is None:
            continue
        if s.dusum_mu:
            miktar = -miktar
        if abs(miktar) < 1e-9:
            continue
        fiyat = fiyatlar.get(s.poz)
        anahtar = (s.poz, s.tanim, s.birim)
        kutu = gruplar.setdefault(
            anahtar,
            {
                "poz": s.poz,
                "tanim": s.tanim,
                "birim": s.birim,
                "miktar": 0.0,
                "fiyat": fiyat,
                "tutar": 0.0,
                "bolum": _bolum(s.poz),
                "dusum": s.dusum_mu,
                "kaynak_eleman": s.eleman_adi,
            },
        )
        kutu["miktar"] += miktar
        if kutu["fiyat"] is None:
            kutu["fiyat"] = fiyat
        kutu["tutar"] = (kutu["fiyat"] or 0.0) * kutu["miktar"]

    bolum_sirasi = {ad: i for i, (ad, _) in enumerate(BOLUMLER)}
    kalemler = sorted(
        gruplar.values(),
        key=lambda k: (bolum_sirasi.get(k["bolum"], 99), k["poz"], k["miktar"]),
    )
    fiyatli = [k for k in kalemler if k["fiyat"] is not None]
    ara_toplam = sum(k["tutar"] for k in fiyatli)
    kdv = ara_toplam * kdv_oran / 100.0
    genel_toplam = ara_toplam + kdv

    fiyatsiz = sorted(
        {k["poz"] for k in kalemler if k["fiyat"] is None}
        | {s.poz for s in sonuc.satirlar if s.poz not in fiyatlar}
    )

    def _oner(var: float | None) -> float | None:
        return var

    oneriler: dict[str, float] = {}
    from hakedis.maliyet import _fiyat_onerisi

    for p in fiyatsiz:
        o = _fiyat_onerisi(fiyatlar, p)
        if o is not None:
            oneriler[p] = o

    for i, k in enumerate(kalemler, 1):
        k["sira"] = i

    return {
        "aktif": True,
        "para_birimi": str(ayarlar.al("maliyet.para_birimi", "TL")),
        "kdv_oran": kdv_oran,
        "beton_sinifi": sinif,
        "kat": sonuc.kat,
        "kaynak_dosya": str(sonuc.kaynak_dosya or ""),
        "kalemler": kalemler,
        "fiyatli_adet": len(fiyatli),
        "toplam_poz": len(kalemler),
        "ara_toplam": ara_toplam,
        "kdv": kdv,
        "genel_toplam": genel_toplam,
        "fiyatsiz_pozlar": fiyatsiz,
        "fiyat_onerileri": oneriler,
        "not": (
            "Birim fiyatlar ORNEKTIR. Kesin bedel icin guncel bakanlik/il "
            "birim fiyatlarini 'Maliyet' bolumune veya birim_fiyatlar.yml "
            "dosyasina girin."
        ),
    }


def kesif_konsol(k: dict, ayrintili: bool = True) -> str:
    """Kesif tablosunu konsola yazdirilabilir metne cevirir."""
    if not k["kalemler"]:
        return "Kesif: metrajda ucretlendirilecek kalem yok."
    satirlar = ["", "ALINAN KESIF  (poz bazinda toplanmis metraj)", "=" * 78]
    satirlar.append(f"Kat: {k['kat'] or '?'}   Betonarme: {k['beton_sinifi']}")
    satirlar.append(
        f"{'NO':>4} {'POZ':<16}{'IMALAT':<34}{'BIRIM':<6}{'MIKTAR':>11}"
        f"  {'TUTAR':>12}"
    )
    satirlar.append("-" * 78)
    son_bolum = None
    bolum_toplam: dict[str, float] = {}
    for kk in k["kalemler"]:
        if kk["bolum"] != son_bolum:
            satirlar.append(f"  {kk['bolum']}")
            son_bolum = kk["bolum"]
        bolum_toplam[son_bolum] = bolum_toplam.get(son_bolum, 0.0) + (
            kk["tutar"] or 0.0
        )
        fiyat = f"{kk['fiyat']:.0f}" if kk["fiyat"] is not None else "  -  "
        tutar = f"{kk['tutar']:>12,.0f}" if kk["fiyat"] is not None else f"{'':>12}"
        satirlar.append(
            f"{kk['sira']:>4} {kk['poz']:<16}{kk['tanim'][:33]:<34}"
            f"{kk['birim']:<6}{kk['miktar']:>11.3f}  {fiyat:>6} {tutar}"
        )
    satirlar.append("-" * 78)
    satirlar.append(f"ARA TOPLAM{'':<68}{k['ara_toplam']:>12,.0f}")
    satirlar.append(f"KDV (%{k['kdv_oran']:g}){'':<67}{k['kdv']:>12,.0f}")
    satirlar.append(
        f"GENEL TOPLAM ({k['para_birimi']}){'':<55}{k['genel_toplam']:>12,.0f}"
    )
    if k["fiyatsiz_pozlar"]:
        ekler = ", ".join(k["fiyatsiz_pozlar"])
        satirlar.append(f"\nFiyat tanimsiz pozlar: {ekler}")
        if k["fiyat_onerileri"]:
            satirlar.append(
                "Onerilen birim fiyatlar: "
                + ", ".join(f"{p}={o:g}" for p, o in k["fiyat_onerileri"].items())
            )
    satirlar.append(f"\n{k['not']}")
    return "\n".join(satirlar)


def _cetvel_yaz(ws, k: dict, baslik: str) -> None:
    """Kesif cetvelini acilmis bir calisma sayfasina yazar."""
    for harf, g in {"A": 6, "B": 14, "C": 48, "D": 8, "E": 14, "F": 14, "G": 16}.items():
        ws.column_dimensions[harf].width = g
    ws["A1"] = baslik
    ws["A1"].font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Kat: {k['kat'] or '?'}").font = Font(bold=True)
    ws.cell(row=3, column=1, value=f"Betonarme sinifi: {k['beton_sinifi']}").font = Font(bold=True)
    ws.cell(row=4, column=1, value=f"Para birimi: {k['para_birimi']}").font = Font(bold=True)

    satir = 6
    basliklar = [
        ("Sira No", "A"), ("Poz No", "B"), ("Imalat (poz tanimi)", "C"),
        ("Birim", "D"), ("Miktar", "E"), ("Birim Fiyat", "F"), ("Tutar", "G"),
    ]
    for i, (ad, _h) in enumerate(basliklar, start=1):
        h = ws.cell(row=satir, column=i, value=ad)
        h.fill = BASLIK_DOLGU
        h.font = BASLIK_YAZI
        h.alignment = Alignment(horizontal="center", vertical="center")
        h.border = CERCEVE
    satir += 1

    son_bolum = None
    bolum_toplam: dict[str, float] = {}
    for kk in k["kalemler"]:
        if kk["bolum"] != son_bolum:
            if son_bolum is not None:
                satir += 1
                ws.cell(row=satir, column=1, value=f"{son_bolum} ARA TOPLAM").font = TOPLAM_YAZI
                ws.cell(
                    row=satir, column=7, value=round(bolum_toplam.get(son_bolum, 0.0), 2)
                ).font = TOPLAM_YAZI
                ws.cell(row=satir, column=7).number_format = "0.00"
                for c in range(1, 8):
                    ws.cell(row=satir, column=c).fill = GRUP_DOLGU
            satir += 1
            son_bolum = kk["bolum"]
            bolum_toplam.setdefault(son_bolum, 0.0)
        bolum_toplam[son_bolum] = bolum_toplam.get(son_bolum, 0.0) + (
            kk["tutar"] or 0.0
        )
        degerler = [
            kk["sira"],
            kk["poz"],
            kk["tanim"],
            kk["birim"],
            kk["miktar"],
            kk["fiyat"] if kk["fiyat"] is not None else "fiyat yok",
            kk["tutar"] if kk["fiyat"] is not None else "-",
        ]
        for i, deger in enumerate(degerler, start=1):
            h = ws.cell(row=satir, column=i, value=deger)
            h.border = CERCEVE
            if i in (5, 6, 7) and isinstance(deger, (int, float)):
                h.number_format = "0.00"
        if kk["dusum"]:
            for c in range(1, 8):
                ws.cell(row=satir, column=c).fill = DUSUM_DOLGU
        satir += 1

    if son_bolum is not None:
        satir += 1
        ws.cell(row=satir, column=1, value=f"{son_bolum} ARA TOPLAM").font = TOPLAM_YAZI
        ws.cell(
            row=satir, column=7, value=round(bolum_toplam.get(son_bolum, 0.0), 2)
        ).font = TOPLAM_YAZI
        ws.cell(row=satir, column=7).number_format = "0.00"
        for c in range(1, 8):
            ws.cell(row=satir, column=c).fill = GRUP_DOLGU

    satir += 1
    for ad, deger in [
        ("ARA TOPLAM", k["ara_toplam"]),
        (f"KDV (%{k['kdv_oran']:g})", k["kdv"]),
        ("GENEL TOPLAM", k["genel_toplam"]),
    ]:
        ws.cell(row=satir, column=1, value=ad).font = TOPLAM_YAZI
        ws.cell(row=satir, column=1).fill = TOPLAM_DOLGU
        h = ws.cell(row=satir, column=7, value=round(deger, 2))
        h.font = TOPLAM_YAZI
        h.number_format = "0.00"
        for c in range(1, 8):
            ws.cell(row=satir, column=c).border = CERCEVE
        satir += 1

    if k["fiyatsiz_pozlar"]:
        satir += 1
        h = ws.cell(
            row=satir, column=1,
            value="Fiyat tanimsiz pozlar (miktar listelendi): "
            + ", ".join(k["fiyatsiz_pozlar"]),
        )
        h.font = UYARI_YAZI
    satir += 1
    ws.cell(row=satir, column=1, value=k["not"]).alignment = Alignment(
        wrap_text=True, vertical="top"
    )


def kesif_sayfasi_yaz(wb: Workbook, k: dict, sayfa: str = "Alinan Kesif") -> None:
    """Var olan workbook'a 'Alinan Kesif' sayfasini ekler."""
    ws = wb.create_sheet(sayfa)
    _cetvel_yaz(ws, k, "ALINAN KESIF  (poz bazinda toplanmis metraj)")


def kesif_excel_yaz(k: dict, hedef: str | Path) -> Path:
    """Kesif tablosunu ayri, ihale teslimine hazir bir Excel dosyasina yazar."""
    from hakedis.report.excel import _rapor_bitir

    hedef = Path(hedef)
    wb = Workbook()
    wb.remove(wb.active)
    kesif_sayfasi_yaz(wb, k)
    _rapor_bitir(wb)
    hedef.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(hedef))
    return hedef


__all__ = [
    "kesif_hesapla",
    "kesif_konsol",
    "kesif_sayfasi_yaz",
    "kesif_excel_yaz",
]