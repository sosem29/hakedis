"""Excel/rapor zenginlestirme: yazdirma (PDF) ayarlari ve filtreler."""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from hakedis.config import ayarlari_yukle
from hakedis.kesif import kesif_excel_yaz, kesif_hesapla
from hakedis.metraj import plandan_metraj
from hakedis.report import excel_yaz

from ornek.ornek_plan_uret import uret


@pytest.fixture(scope="module")
def plan(tmp_path_factory) -> Path:
    return uret(tmp_path_factory.mktemp("rapor") / "kalip_plani.dxf")


@pytest.fixture(scope="module")
def workbook(plan, tmp_path_factory) -> Path:
    ay = ayarlari_yukle()
    s, _ = plandan_metraj(str(plan), ay)
    yol = tmp_path_factory.mktemp("cikti") / "metraj.xlsx"
    excel_yaz(s, yol, ayarlar=ay)
    return yol


class TestYazdirmaAyarlari:
    def test_uzun_sayfalar_yatay_a4(self, workbook):
        wb = load_workbook(workbook)
        ws = wb["Metraj Cetveli"]
        assert ws.page_setup.orientation == "landscape"
        assert ws.page_setup.paperSize == 9  # A4
        assert ws.page_setup.fitToWidth == 1

    def test_kisa_sayfa_dikey(self, workbook):
        wb = load_workbook(workbook)
        ws = wb["Uyarilar"]
        assert ws.page_setup.orientation == "portrait"

    def test_baslik_her_sayfada_tekrarlanir(self, workbook):
        wb = load_workbook(workbook)
        ws = wb["Metraj Cetveli"]
        assert ws.print_title_rows == "$1:$1"

    def test_altbilgi_sayfa_numarasi(self, workbook):
        wb = load_workbook(workbook)
        ws = wb["Metraj Cetveli"]
        assert "&P" in ws.oddFooter.center.text

    def test_otomatik_filtre(self, workbook):
        wb = load_workbook(workbook)
        cetvel = wb["Metraj Cetveli"]
        assert cetvel.auto_filter.ref.startswith("A1:")
        kesif = wb["Alinan Kesif"]
        assert kesif.auto_filter.ref.startswith("A6:")

    def test_kesif_ayri_dosya_aynı_ayarlar(self, workbook, plan, tmp_path):
        ay = ayarlari_yukle()
        s, _ = plandan_metraj(str(plan), ay)
        yol = kesif_excel_yaz(kesif_hesapla(s, ay), tmp_path / "kesif.xlsx")
        wb = load_workbook(yol)
        ws = wb["Alinan Kesif"]
        assert ws.page_setup.orientation == "landscape"
        assert ws.auto_filter.ref.startswith("A6:")


class TestOzetZenginlestirme:
    def test_uretım_tarihi(self, workbook):
        wb = load_workbook(workbook)
        ws = wb["Ozet"]
        degerler = {
            r[0]: r[1]
            for r in ws.iter_rows(min_row=1, values_only=True)
            if r[0]
        }
        assert "Uretim tarihi" in degerler

    def test_sayi_formatlari(self, workbook):
        wb = load_workbook(workbook)
        ws = wb["Ozet"]
        uygun = False
        for row in ws.iter_rows(min_row=1, max_row=15):
            if row[0].value == "Kolon":
                uygun = True
                assert row[2].number_format == "0.000"  # beton
                assert row[4].number_format == "0.0"  # demir
        assert uygun
