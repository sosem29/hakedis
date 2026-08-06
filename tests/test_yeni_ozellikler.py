"""Yeni ozellikler: yaklasik demir, merdiven egim katsayisi, guseli/mantar
doseme ve cok katli/paftali (toplu) calisma testleri."""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from hakedis.config import ayarlari_yukle
from hakedis.metraj import metraj_hesapla, plandan_metraj, sonuclari_birlestir
from hakedis.model import Eleman, ElemanTipi, Nokta
from hakedis.report import (
    excel_yaz_toplu,
    konsol_ozeti,
    konsol_ozeti_toplu,
)

from ornek.ornek_plan_uret import uret


def dikdortgen(x0, y0, x1, y1) -> list[Nokta]:
    return [Nokta(x0, y0), Nokta(x1, y0), Nokta(x1, y1), Nokta(x0, y1)]


@pytest.fixture(scope="module")
def plan(tmp_path_factory) -> Path:
    return uret(tmp_path_factory.mktemp("plan") / "kalip_plani.dxf")


def _doseme_elemani() -> Eleman:
    e = Eleman(
        ad="D01",
        tip=ElemanTipi.DOSEME,
        kat="Kat",
        cevre=dikdortgen(0, 0, 10, 10),
    )
    e.olculer["t"] = 0.15
    return e


class TestYaklasikDemir:
    def test_varsayilan_kapali(self, plan):
        s, _ = plandan_metraj(str(plan), ayarlari_yukle())
        assert all(r.kg is None for r in s.satirlar)
        assert not any(k["demir_kg"] for k in s.ozet().values())

    def test_aktifte_kilo_satirlari(self, plan):
        ayarlar = ayarlari_yukle()
        ayarlar.ham["donati"] = {"aktif": True, "katsayilar": {"kolon": 110}}
        s, _ = plandan_metraj(str(plan), ayarlar)
        kg_satirlar = [r for r in s.satirlar if r.birim == "kg"]
        assert kg_satirlar, "donati aktifken kg satirlari uretilmeli"
        # 6 kolon x (0.18 m3 x 2.85) x 110
        beklenen = 6 * (0.30 * 0.60 * 2.85) * 110
        assert s.ozet()["Kolon"]["demir_kg"] == pytest.approx(beklenen, abs=0.1)
        for r in kg_satirlar:
            assert r.kg == pytest.approx(r.hacim * 110, abs=1e-6)
        assert any("YAKLASIK" in u for u in s.uyarilar)

    def test_sifir_katsayi_satir_uretmez(self):
        e = Eleman(
            ad="S01",
            tip=ElemanTipi.KOLON,
            kat="Kat",
            cevre=dikdortgen(0, 0, 0.3, 0.6),
        )
        ayarlar = ayarlari_yukle()
        ayarlar.ham["donati"] = {
            "aktif": True,
            "katsayilar": {"kolon": 0},
        }
        s = metraj_hesapla([e], ayarlar)
        assert not any(r.birim == "kg" for r in s.satirlar)

    def test_miktar_property_kg_verir(self, plan):
        ayarlar = ayarlari_yukle()
        ayarlar.ham["donati"] = {"aktif": True, "katsayilar": {"kolon": 100}}
        s, _ = plandan_metraj(str(plan), ayarlar)
        kg_row = next(r for r in s.satirlar if r.birim == "kg")
        assert kg_row.miktar == pytest.approx(kg_row.kg)


class TestMerdivenEgim:
    def _merdiven(self):
        e = Eleman(
            ad="M01",
            tip=ElemanTipi.MERDIVEN,
            kat="Kat",
            cevre=dikdortgen(0, 0, 4, 3),
        )
        e.olculer["t"] = 0.14
        return e

    def test_varsayilan_riht_basamaktan_hesaplanir(self):
        # k = sqrt(1 + (0.175/0.28)^2)
        k = (1 + (0.175 / 0.28) ** 2) ** 0.5
        s = metraj_hesapla([self._merdiven()], ayarlari_yukle())
        beton = next(r for r in s.satirlar if r.tip == ElemanTipi.MERDIVEN and r.birim == "m3")
        assert beton.hacim == pytest.approx(12 * k * 0.14, abs=1e-3)
        assert "k=" in beton.tanim
        kalip = next(r for r in s.satirlar if r.tip == ElemanTipi.MERDIVEN and r.birim == "m2")
        assert kalip.alan == pytest.approx(12 * k, abs=1e-3)

    def test_dogrudan_katsayi_onceliklidir(self):
        ayarlar = ayarlari_yukle()
        ayarlar.ham["merdiven"] = {"egim_katsayisi": 1.5}
        s = metraj_hesapla([self._merdiven()], ayarlar)
        beton = next(r for r in s.satirlar if r.birim == "m3")
        assert beton.hacim == pytest.approx(12 * 1.5 * 0.14, abs=1e-3)

    def test_katsayi_yoksa_plan_izdusumu(self):
        ayarlar = ayarlari_yukle()
        ayarlar.ham["merdiven"] = {"riht": 0, "basamak": 0}
        s = metraj_hesapla([self._merdiven()], ayarlar)
        beton = next(r for r in s.satirlar if r.birim == "m3")
        assert beton.hacim == pytest.approx(12 * 0.14, abs=1e-6)
        assert "k=1.00" in beton.tanim


class TestOzelDoseme:
    def test_guseli_hacim_katsayisi(self):
        ayarlar = ayarlari_yukle()
        ayarlar.ham["doseme"] = {"tip": "guseli", "guseli_hacim_katsayisi": 1.35}
        s = metraj_hesapla([_doseme_elemani()], ayarlar)
        beton = next(r for r in s.satirlar if r.tip == ElemanTipi.DOSEME and r.birim == "m3")
        assert beton.hacim == pytest.approx(100 * 0.15 * 1.35, abs=1e-3)
        assert any("GUSELI" in d for d in beton.detay)
        assert any("guseli" in u for u in s.uyarilar)

    def test_mantar_kolon_ustu_ilavesi(self):
        ayarlar = ayarlari_yukle()
        ayarlar.ham["doseme"] = {
            "tip": "mantar",
            "mantar_kolon_ustu_artisi": 0.05,
            "mantar_kolon_baslik_alani": 1.00,
        }
        kolonlar = [
            Eleman(
                ad="S01", tip=ElemanTipi.KOLON, kat="Kat",
                cevre=dikdortgen(2.5, 2.5, 3.0, 3.0),
            ),
            Eleman(
                ad="S02", tip=ElemanTipi.KOLON, kat="Kat",
                cevre=dikdortgen(7.0, 7.0, 7.5, 7.5),
            ),
        ]
        s = metraj_hesapla([_doseme_elemani()] + kolonlar, ayarlar)
        ilave = next(r for r in s.satirlar if "Mantar" in r.tanim)
        assert ilave.hacim == pytest.approx(2 * 1.00 * 0.05, abs=1e-6)
        # Ana doseme betonu ilavesiz kaliyor
        beton = next(
            r for r in s.satirlar
            if r.tip == ElemanTipi.DOSEME and r.birim == "m3" and r.hacim == pytest.approx(15.0)
        )
        assert beton is not None

    def test_normal_tip_degisiklik_yapmaz(self):
        s = metraj_hesapla([_doseme_elemani()], ayarlari_yukle())
        beton = next(r for r in s.satirlar if r.birim == "m3")
        assert beton.hacim == pytest.approx(100 * 0.15, abs=1e-6)


class TestToplu:
    def _iki_kat(self, plan) -> list:
        s1, _ = plandan_metraj(
            str(plan), ayarlari_yukle().guncelle(kat_adi="Giris")
        )
        s2, _ = plandan_metraj(
            str(plan), ayarlari_yukle().guncelle(kat_adi="1. Kat")
        )
        return [s1, s2]

    def test_birlestirici_toplamlar(self, plan):
        s1, s2 = self._iki_kat(plan)
        b = sonuclari_birlestir([s1, s2])
        assert b.ozet()["Kolon"]["beton_m3"] == pytest.approx(
            2 * s1.ozet()["Kolon"]["beton_m3"], abs=1e-6
        )
        assert b.ozet()["Kolon"]["adet"] == pytest.approx(
            2 * s1.ozet()["Kolon"]["adet"]
        )
        for u in b.uyarilar:
            assert u.startswith("[Giris]") or u.startswith("[1. Kat]")

    def test_birlestirici_uyari_oneki(self):
        from hakedis.model import MetrajSonucu

        m1 = MetrajSonucu(kat="Giris", uyarilar=["kolon dusuk guven"])
        m2 = MetrajSonucu(kat="1. Kat", uyarilar=["kiris net aciklik"])
        b = sonuclari_birlestir([m1, m2])
        assert b.uyarilar == ["[Giris] kolon dusuk guven", "[1. Kat] kiris net aciklik"]

    def test_excel_toplu_yazilir(self, plan, tmp_path):
        yol = excel_yaz_toplu(self._iki_kat(plan), tmp_path / "toplu.xlsx")
        from openpyxl import load_workbook

        wb = load_workbook(yol)
        assert "Kat Ozeti" in wb.sheetnames
        ws = wb["Kat Ozeti"]
        degerler = [r[0] for r in ws.iter_rows(min_row=5, values_only=True) if r[0]]
        assert "Giris" in degerler and "1. Kat" in degerler

    def test_konsol_toplu_ozeti(self, plan):
        metin = konsol_ozeti_toplu(self._iki_kat(plan))
        assert "TOPLU KIRIK OLCU METRAJI" in metin
        assert "Giris" in metin and "1. Kat" in metin
        assert "TOPLAM" in metin

    def test_bos_liste_guvenli(self):
        b = sonuclari_birlestir([])
        assert b.ozet() == {}

    def test_toplu_cli(self, plan, tmp_path, capsys):
        from hakedis.cli import main

        cikti = tmp_path / "toplu.xlsx"
        kod = main(
            [
                "toplu",
                str(plan),
                str(plan),
                "--kat-adlari",
                "Giris",
                "1. Kat",
                "--cikti",
                str(cikti),
            ]
        )
        assert kod == 0
        assert cikti.exists()
        out = capsys.readouterr().out
        assert "TOPLU KIRIK OLCU METRAJI" in out
        assert "Giris" in out and "1. Kat" in out


class TestCiktiTutarliligi:
    def test_konsol_demir_sutunu(self, plan):
        ayarlar = ayarlari_yukle()
        ayarlar.ham["donati"] = {"aktif": True, "katsayilar": {"kolon": 110}}
        s, _ = plandan_metraj(str(plan), ayarlar)
        metin = konsol_ozeti(s)
        assert "Demir (kg)" in metin

    def test_demir_detay_satirda(self, plan):
        ayarlar = ayarlari_yukle()
        ayarlar.ham["donati"] = {"aktif": True, "katsayilar": {"kolon": 110}}
        s, _ = plandan_metraj(str(plan), ayarlar)
        kg = next(r for r in s.satirlar if r.birim == "kg")
        assert any("YAKLASIK" in d for d in kg.detay)
