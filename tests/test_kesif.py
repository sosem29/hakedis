"""Alinan kesif (poz gruplu metraj + birim fiyat) testleri."""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from hakedis.config import ayarlari_yukle
from hakedis.kesif import (
    kesif_excel_yaz,
    kesif_hesapla,
    kesif_konsol,
    kesif_sayfasi_yaz,
)
from hakedis.metraj import plandan_metraj
from hakedis.model import ElemanTipi, KirikOlcuSatiri, MetrajSonucu

from ornek.ornek_plan_uret import uret


@pytest.fixture(scope="module")
def plan(tmp_path_factory) -> Path:
    return uret(tmp_path_factory.mktemp("kesif") / "kalip_plani.dxf")


@pytest.fixture(scope="module")
def kesif(plan) -> dict:
    ayarlar = ayarlari_yukle()
    ayarlar.ham.setdefault("maliyet", {})["aktif"] = True
    s, _ = plandan_metraj(str(plan), ayarlar)
    return kesif_hesapla(s, ayarlar)


class TestKesifHesabi:
    def test_pozlar_tekil_ve_toplanmis(self, kesif, plan):
        ayarlar = ayarlari_yukle()
        s, _ = plandan_metraj(str(plan), ayarlar)
        # poz bazinda tek satir: 16.058/1-K bir kez, miktari 6 kolonun toplami
        kolon_beton = [k for k in kesif["kalemler"] if k["poz"] == "16.058/1-K"]
        assert len(kolon_beton) == 1
        oz = s.ozet()["Kolon"]
        assert kolon_beton[0]["miktar"] == pytest.approx(oz["beton_m3"], abs=1e-3)
        # ayni pozdaki kalip satirlari (dusum ayri imalat olarak toplanir)
        kalip_k = [k for k in kesif["kalemler"] if k["poz"] == "21.011/K"]
        assert sum(k["miktar"] for k in kalip_k) == pytest.approx(
            oz["kalip_m2"], abs=1e-3
        )

    def test_tutarlar_ozetle_uyumlu(self, kesif):
        # beton kalemlerinin toplami = genel beton miktari
        beton = sum(k["miktar"] for k in kesif["kalemler"] if k["birim"] == "m3")
        assert beton == pytest.approx(18.933, abs=1e-3)
        # her kalem tutari = miktar x fiyat
        for k in kesif["kalemler"]:
            if k["fiyat"] is not None:
                assert k["tutar"] == pytest.approx(k["miktar"] * k["fiyat"], rel=1e-6)

    def test_toplama_uyari_ve_bolumler(self, kesif):
        assert kesif["ara_toplam"] < kesif["genel_toplam"]
        assert kesif["kdv"] == pytest.approx(kesif["ara_toplam"] * 0.20)
        bolumler = {k["bolum"] for k in kesif["kalemler"]}
        assert bolumler and all(b in {"BETONARME", "SIVA-BADANA", "DOGRAMA",
                                      "DOSEME KAPLAMA", "DIGER"} for b in bolumler)

    def test_fiyatsiz_poz_ayri_listelenir(self):
        satirlar = [
            KirikOlcuSatiri(
                poz="16.058/1-K",
                eleman_adi="S01",
                tip=ElemanTipi.KOLON,
                tanim="Kolon betonu",
                benzer=1,
                hacim=1.0,
                birim="m3",
                kat="Kat",
            ),
            KirikOlcuSatiri(
                poz="99.999",
                eleman_adi="X01",
                tip=ElemanTipi.BILINMEYEN,
                tanim="Fiyatsiz imalat",
                benzer=1,
                alan=5.0,
                birim="m2",
                kat="Kat",
            ),
        ]
        sonuc = MetrajSonucu(kat="Kat", satirlar=satirlar)
        k = kesif_hesapla(sonuc, ayarlari_yukle())
        assert "99.999" in k["fiyatsiz_pozlar"]
        fiyatsiz = [x for x in k["kalemler"] if x["poz"] == "99.999"]
        assert len(fiyatsiz) == 1
        assert fiyatsiz[0]["fiyat"] is None
        # fiyatli kalemin tutari hesaba dahil
        assert k["ara_toplam"] > 0

    def test_dusum_miktar_negatif(self):
        satirlar = [
            KirikOlcuSatiri(
                poz="21.011/K",
                eleman_adi="K01",
                tip=ElemanTipi.KIRIS,
                tanim="Kiris kalibi",
                benzer=1,
                alan=1.0,
                birim="m2",
                kat="Kat",
            ),
            KirikOlcuSatiri(
                poz="21.011/K",
                eleman_adi="K01",
                tip=ElemanTipi.KIRIS,
                tanim="Kiris kalibi",
                benzer=1,
                alan=0.2,
                birim="m2",
                kat="Kat",
                dusum_mu=True,
            ),
        ]
        sonuc = MetrajSonucu(kat="Kat", satirlar=satirlar)
        k = kesif_hesapla(sonuc, ayarlari_yukle())
        satir = [x for x in k["kalemler"] if x["poz"] == "21.011/K"][0]
        assert satir["miktar"] == pytest.approx(0.8)
        assert satir["tutar"] == pytest.approx(0.8 * 480)


class TestKesifCiktilari:
    def test_konsol_ciktisi(self, kesif):
        metin = kesif_konsol(kesif)
        assert "ALINAN KESIF" in metin
        assert "GENEL TOPLAM" in metin and "ARA TOPLAM" in metin
        assert "16.058/1-K" in metin

    def test_excel_yazilir(self, kesif, tmp_path):
        hedef = kesif_excel_yaz(kesif, tmp_path / "kesif.xlsx")
        assert hedef.exists()
        from openpyxl import load_workbook

        wb = load_workbook(hedef)
        assert "Alinan Kesif" in wb.sheetnames
        ws = wb["Alinan Kesif"]
        baslik = ws["A1"].value
        assert "ALINAN KESIF" in baslik
        # toplam satirlari yazildi
        degerler = " ".join(
            str(r[0]) for r in ws.iter_rows(min_row=1, values_only=True) if r[0]
        )
        assert "GENEL TOPLAM" in degerler

    def test_metraj_excelinde_kesif_sayfasi(self, plan, tmp_path):
        from hakedis.report import excel_yaz

        ayarlar = ayarlari_yukle()
        s, _ = plandan_metraj(str(plan), ayarlar)
        hedef = excel_yaz(s, tmp_path / "metraj.xlsx", ayarlar=ayarlar)
        from openpyxl import load_workbook

        wb = load_workbook(hedef)
        assert "Alinan Kesif" in wb.sheetnames


class TestKesifCli:
    def test_plandan_kesif(self, plan, tmp_path, capsys):
        from hakedis.cli import main

        cikti = tmp_path / "kesif.xlsx"
        assert main(["kesif", str(plan), "--cikti", str(cikti)]) == 0
        assert cikti.exists()
        out = capsys.readouterr().out
        assert "ALINAN KESIF" in out and "GENEL TOPLAM" in out

    def test_json_kesif(self, plan, tmp_path, capsys):
        from hakedis.cli import main

        mj = tmp_path / "m.json"
        assert main(["metraj", str(plan), "--json", str(mj), "--svg-yok"]) == 0
        cikti = tmp_path / "kesif2.xlsx"
        assert main(["kesif", str(mj), "--cikti", str(cikti)]) == 0
        assert cikti.exists()
        out = capsys.readouterr().out
        assert "ALINAN KESIF" in out
