"""Uctan uca testler: DXF ve PDF planlarindan metraj uretimi.

Beklenen degerler elle hesaplanmistir; ornek plandaki geometri sabittir.
Bu testler metraj formullerinin sessizce degismesini engeller.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from hakedis.config import ayarlari_yukle
from hakedis.metraj import plandan_metraj
from hakedis.model import ElemanTipi
from hakedis.report import excel_yaz, konsol_ozeti, svg_yaz
from ornek.ornek_plan_uret import uret
from tests.yardimci import kalip_plani_pdf

# Ornek plan sabitleri (bkz. ornek/ornek_plan_uret.py)
KAT_YUKSEKLIGI = 3.00
DOSEME_KALINLIGI = 0.15
NET_YUKSEKLIK = KAT_YUKSEKLIGI - DOSEME_KALINLIGI  # 2.85


@pytest.fixture(scope="module")
def plan(tmp_path_factory) -> Path:
    return uret(tmp_path_factory.mktemp("plan") / "kalip_plani.dxf")


@pytest.fixture(scope="module")
def sonuc(plan):
    s, _ = plandan_metraj(str(plan), ayarlari_yukle())
    return s


class TestElemanTespiti:
    def test_eleman_sayilari(self, sonuc):
        assert len(sonuc.tipe_gore(ElemanTipi.KOLON)) == 6
        assert len(sonuc.tipe_gore(ElemanTipi.PERDE)) == 1
        assert len(sonuc.tipe_gore(ElemanTipi.KIRIS)) == 5
        assert len(sonuc.tipe_gore(ElemanTipi.DOSEME)) == 1

    def test_kolon_kesiti_cizimden_okunur(self, sonuc):
        for k in sonuc.tipe_gore(ElemanTipi.KOLON):
            assert k.olculer["b"] == pytest.approx(0.30, abs=1e-6)
            assert k.olculer["h"] == pytest.approx(0.60, abs=1e-6)

    def test_kolon_adlari_etiketten_gelir(self, sonuc):
        adlar = {k.ad for k in sonuc.tipe_gore(ElemanTipi.KOLON)}
        assert adlar == {f"S{i:02d}" for i in range(1, 7)}
        assert not any("*" in a for a in adlar)

    def test_l_perde_kirik_olcusu(self, sonuc):
        p = sonuc.tipe_gore(ElemanTipi.PERDE)[0]
        assert p.olculer["t"] == pytest.approx(0.25, abs=1e-6)
        assert len(p.segmentler) == 2, "L perde iki kirik olcu parcasi vermeli"
        assert p.toplam_uzunluk == pytest.approx(3.50, abs=1e-6)
        assert p.olculer["serbest_uc"] == 2

    def test_kiris_net_aciklik(self, sonuc):
        """Boyuna kirisler 12.30 brutten 3x0.30 kolon dusulerek 11.40 olmali."""
        uzunlar = [
            k
            for k in sonuc.tipe_gore(ElemanTipi.KIRIS)
            if k.olculer.get("brut_uzunluk", 0) > 10
        ]
        assert len(uzunlar) == 2
        for k in uzunlar:
            assert k.olculer["brut_uzunluk"] == pytest.approx(12.30, abs=1e-6)
            assert k.olculer["net_uzunluk"] == pytest.approx(11.40, abs=0.02)
            assert len(k.segmentler) == 2, "iki aciklik olmali"

    def test_doseme_boslugu_dusulur(self, sonuc):
        d = sonuc.tipe_gore(ElemanTipi.DOSEME)[0]
        assert d.olculer["brut_alan"] == pytest.approx(68.88, abs=1e-4)
        assert d.olculer["bosluk_alani"] == pytest.approx(0.80, abs=1e-4)
        assert d.olculer["net_alan"] == pytest.approx(68.08, abs=1e-4)
        assert d.olculer["t"] == pytest.approx(0.15), "TD=15 etiketinden okunmali"


class TestMetrajDegerleri:
    def test_kolon_betonu(self, sonuc):
        # 6 x (0.30 x 0.60) x 2.85
        beklenen = 6 * 0.30 * 0.60 * NET_YUKSEKLIK
        assert sonuc.ozet()["Kolon"]["beton_m3"] == pytest.approx(beklenen, abs=1e-3)

    def test_kolon_kalibi_kiris_dusumuyle(self, sonuc):
        # Brut: 6 x 1.80 x 2.85 = 30.78
        brut = 6 * 1.80 * NET_YUKSEKLIK
        # Dusum: her kolonda saplanan kiris yuzleri x 0.25 x (0.50-0.15)
        # Kose kolonlar (4 adet): 2 yuz | ara kolonlar (2 adet): 3 yuz
        dusum = (4 * 2 + 2 * 3) * 0.25 * 0.35
        assert sonuc.ozet()["Kolon"]["kalip_m2"] == pytest.approx(
            brut - dusum, abs=1e-3
        )

    def test_perde_betonu(self, sonuc):
        assert sonuc.ozet()["Perde"]["beton_m3"] == pytest.approx(
            3.50 * 0.25 * NET_YUKSEKLIK, abs=1e-3
        )

    def test_perde_kalibi_iki_yuz_arti_bas(self, sonuc):
        beklenen = 2 * 3.50 * NET_YUKSEKLIK + 2 * 0.25 * NET_YUKSEKLIK
        assert sonuc.ozet()["Perde"]["kalip_m2"] == pytest.approx(beklenen, abs=1e-3)

    def test_kiris_betonu_dosemesiz_yukseklikle(self, sonuc):
        # 2 x 11.40 + 3 x 4.40 = 36.00 m net boy, kesit 0.25 x (0.50-0.15)
        beklenen = 36.00 * 0.25 * 0.35
        assert sonuc.ozet()["Kiris"]["beton_m3"] == pytest.approx(beklenen, abs=1e-2)

    def test_doseme_betonu(self, sonuc):
        assert sonuc.ozet()["Doseme"]["beton_m3"] == pytest.approx(
            68.08 * 0.15, abs=1e-3
        )

    def test_doseme_kalibi_cifte_dusum_yapmaz(self, sonuc):
        """Kiris + kolon + perde ayak izleri BIRLESIM olarak dusulmeli."""
        kiris_izi = 36.00 * 0.25  # 9.00
        kolon_izi = 6 * 0.30 * 0.60  # 1.08
        perde_izi = 0.875
        beklenen = 68.08 - (kiris_izi + kolon_izi + perde_izi)
        assert sonuc.ozet()["Doseme"]["kalip_m2"] == pytest.approx(beklenen, abs=1e-2)

    def test_kat_yuksekligi_degisimi_metraji_etkiler(self, plan):
        ayarlar = ayarlari_yukle().guncelle(kat_yuksekligi=3.50)
        s, _ = plandan_metraj(str(plan), ayarlar)
        assert s.ozet()["Kolon"]["beton_m3"] == pytest.approx(
            6 * 0.18 * (3.50 - 0.15), abs=1e-3
        )


class TestKirikOlcuKirilimi:
    def test_her_satirda_formul_ve_detay_var(self, sonuc):
        for s in sonuc.satirlar:
            assert s.formul or s.dusum_mu, f"{s.eleman_adi}: formul bos"
            assert s.detay, f"{s.eleman_adi}: kirik olcu detayi bos"

    def test_perde_detayi_koordinat_tasir(self, sonuc):
        p = [s for s in sonuc.satirlar if s.tip == ElemanTipi.PERDE][0]
        assert any("->" in d and "=" in d for d in p.detay)

    def test_doseme_detayi_kose_koordinatlari(self, sonuc):
        d = [s for s in sonuc.satirlar if s.tip == ElemanTipi.DOSEME][0]
        assert any("kose" in x for x in d.detay)

    def test_toplamlar_satirlarla_tutarli(self, sonuc):
        beton = sum(
            (-1 if s.dusum_mu else 1) * (s.hacim or 0)
            for s in sonuc.satirlar
            if s.birim == "m3"
        )
        ozet_beton = sum(v["beton_m3"] for v in sonuc.ozet().values())
        assert beton == pytest.approx(ozet_beton, abs=1e-6)


class TestCiktilar:
    def test_excel_yazilir(self, sonuc, tmp_path):
        from openpyxl import load_workbook

        yol = excel_yaz(sonuc, tmp_path / "metraj.xlsx")
        assert yol.exists() and yol.stat().st_size > 0
        wb = load_workbook(yol)
        assert {"Ozet", "Metraj Cetveli", "Kirik Olcu", "Elemanlar", "Uyarilar"} <= set(
            wb.sheetnames
        )

    def test_svg_yazilir(self, sonuc, tmp_path):
        yol = svg_yaz(sonuc, tmp_path / "kontrol.svg")
        icerik = yol.read_text(encoding="utf-8")
        assert icerik.startswith("<svg")
        assert icerik.rstrip().endswith("</svg>")
        for e in sonuc.elemanlar:
            assert e.ad.replace("*", "") in icerik

    def test_konsol_ozeti(self, sonuc):
        metin = konsol_ozeti(sonuc, ayrintili=True)
        assert "KIRIK OLCU METRAJI" in metin
        assert "TOPLAM" in metin


@pytest.fixture
def pdf_ayarlari():
    ayarlar = ayarlari_yukle()
    # 1 pt = 1 cm olacak sekilde kalibre et
    ayarlar.ham["pdf"]["kalibrasyon"] = {"pdf_mesafe": 1.0, "gercek_mesafe": 0.01}
    ayarlar.ham["pdf"]["renk_esleme"] = {
        "#ff0000": "kolon",
        "#0000ff": "kiris",
        "#999999": "doseme",
    }
    return ayarlar


class TestPdfYolu:
    def test_pdf_kolonlari_okunur(self, tmp_path, pdf_ayarlari):
        pdf = kalip_plani_pdf(tmp_path / "plan.pdf")
        sonuc, cizim = plandan_metraj(str(pdf), pdf_ayarlari)
        kolonlar = sonuc.tipe_gore(ElemanTipi.KOLON)
        assert len(kolonlar) == 6
        for k in kolonlar:
            assert k.olculer["b"] == pytest.approx(0.30, abs=0.005)
            assert k.olculer["h"] == pytest.approx(0.60, abs=0.005)

    def test_pdf_kirisleri_eslesir(self, tmp_path, pdf_ayarlari):
        pdf = kalip_plani_pdf(tmp_path / "plan2.pdf")
        sonuc, _ = plandan_metraj(str(pdf), pdf_ayarlari)
        kirisler = sonuc.tipe_gore(ElemanTipi.KIRIS)
        assert len(kirisler) == 5
        for k in kirisler:
            assert k.olculer["b"] == pytest.approx(0.25, abs=0.005)

    def test_olcek_yanlissa_olculer_orantili_kayar(self, tmp_path):
        """Olcek iki katina cikarsa tum uzunluklar iki katina cikmali."""
        pdf = kalip_plani_pdf(tmp_path / "plan3.pdf")
        sonuclar = []
        for gercek in (0.01, 0.02):
            ayarlar = ayarlari_yukle()
            ayarlar.ham["pdf"]["kalibrasyon"] = {
                "pdf_mesafe": 1.0,
                "gercek_mesafe": gercek,
            }
            ayarlar.ham["pdf"]["renk_esleme"] = {"#ff0000": "kolon"}
            s, _ = plandan_metraj(str(pdf), ayarlar)
            sonuclar.append(s.tipe_gore(ElemanTipi.KOLON)[0].olculer["b"])
        assert sonuclar[1] == pytest.approx(sonuclar[0] * 2, abs=1e-6)

    def test_vektorsuz_pdf_acik_hata_verir(self, tmp_path):
        from tests.yardimci import basit_pdf_yaz

        bos = basit_pdf_yaz(tmp_path / "bos.pdf", "BT /F1 12 Tf 100 700 Td (x) Tj ET")
        with pytest.raises(ValueError, match="taranmis"):
            plandan_metraj(str(bos), ayarlari_yukle())


YABANCI_ADLAR = {
    "KOLON": "A$C-DIKME-01",
    "PERDE": "BA_WALL_X",
    "KIRIS": "ZZ-BEAM-LAYER",
    "DOSEME": "PLATE_OUTLINE",
    "BOSLUK": "VOID_1",
    "YAZI": "ANNO-TXT",
    "AKS": "GRID_REF",
}


@pytest.fixture(scope="module")
def yabanci_plan(plan, tmp_path_factory) -> Path:
    """Ayni plani, hicbir desene uymayan katman adlariyla yeniden yazar."""
    import ezdxf

    doc = ezdxf.readfile(str(plan))
    for kat in list(doc.layers):
        if kat.dxf.name in YABANCI_ADLAR:
            kat.dxf.name = YABANCI_ADLAR[kat.dxf.name]
    for e in doc.modelspace():
        if e.dxf.layer in YABANCI_ADLAR:
            e.dxf.layer = YABANCI_ADLAR[e.dxf.layer]
    yol = tmp_path_factory.mktemp("yabanci") / "yabanci.dxf"
    doc.saveas(str(yol))
    return yol


class TestOtomatikKesif:
    """Katman adlari taninmadiginda geometriden siniflandirma.

    Katman adi her ofiste farklidir; sistem bunu kullaniciya sordurmadan
    cizimin kendisinden cikarabilmelidir.
    """

    def test_tum_katmanlar_dogru_siniflanir(self, yabanci_plan):
        from hakedis.otomatik import imzalari_cikar
        from hakedis.readers import cizim_oku

        ayarlar = ayarlari_yukle()
        imzalar = imzalari_cikar(cizim_oku(str(yabanci_plan), ayarlar), ayarlar)
        bulunan = {i.katman: i.onerilen_tip for i in imzalar}
        assert bulunan == {
            "A$C-DIKME-01": "kolon",
            "BA_WALL_X": "perde",
            "ZZ-BEAM-LAYER": "kiris",
            "PLATE_OUTLINE": "doseme",
            "VOID_1": "bosluk",
            "ANNO-TXT": "metin",
            "GRID_REF": "yoksay",
        }

    def test_l_perde_kalinliktan_taninir(self, yabanci_plan):
        """L kesitte sinir dikdortgeninin kisa kenari kalinlik degildir."""
        from hakedis.otomatik import imzalari_cikar
        from hakedis.readers import cizim_oku

        ayarlar = ayarlari_yukle()
        imzalar = imzalari_cikar(cizim_oku(str(yabanci_plan), ayarlar), ayarlar)
        perde = next(i for i in imzalar if i.katman == "BA_WALL_X")
        assert perde.ortanca_en == pytest.approx(0.25, abs=1e-6)

    def test_bosluk_kolondan_ayrilir(self, yabanci_plan):
        """Cevresinde kiris ucu olmayan kapali alan kolon degil bosluktur."""
        from hakedis.otomatik import imzalari_cikar
        from hakedis.readers import cizim_oku

        ayarlar = ayarlari_yukle()
        imzalar = imzalari_cikar(cizim_oku(str(yabanci_plan), ayarlar), ayarlar)
        assert next(i for i in imzalar if i.katman == "VOID_1").onerilen_tip == "bosluk"

    def test_metraj_ayni_sonucu_verir(self, sonuc, yabanci_plan):
        """Sifir yapilandirmayla, taninan katmanlarla ayni metraj cikmali."""
        yabanci, _ = plandan_metraj(str(yabanci_plan), ayarlari_yukle())
        for tip, degerler in sonuc.ozet().items():
            for anahtar, beklenen in degerler.items():
                assert yabanci.ozet()[tip][anahtar] == pytest.approx(
                    beklenen, abs=1e-6
                ), f"{tip}/{anahtar} tutmadi"

    def test_otomatik_elemanlar_dusuk_guvenle_isaretlenir(self, yabanci_plan):
        yabanci, _ = plandan_metraj(str(yabanci_plan), ayarlari_yukle())
        assert all(e.guven < 0.7 for e in yabanci.elemanlar)
        assert any("otomatik" in u for u in yabanci.uyarilar)

    def test_yapilandirma_yazilabilir(self, yabanci_plan, tmp_path):
        import yaml

        from hakedis.otomatik import imzalari_cikar, yapilandirma_metni
        from hakedis.readers import cizim_oku

        ayarlar = ayarlari_yukle()
        imzalar = imzalari_cikar(cizim_oku(str(yabanci_plan), ayarlar), ayarlar)
        metin = yapilandirma_metni(imzalar)
        veri = yaml.safe_load(metin)
        assert set(veri["katmanlar"]) >= {"kolon", "perde", "kiris", "doseme"}

        # Asil olcut: uretilen desenler gercekten o katmanlari yakalamali.
        # Katman adlari regex olarak kacislanir ('A$C-...' -> 'A\$C\-...'),
        # bu yuzden metne degil davranisa bakilir.
        yeni = ayarlari_yukle()
        yeni.ham["katmanlar"] = veri["katmanlar"]
        assert yeni.katman_tipi("A$C-DIKME-01") == "kolon"
        assert yeni.katman_tipi("BA_WALL_X") == "perde"
        assert yeni.katman_tipi("ZZ-BEAM-LAYER") == "kiris"
        # Benzer ama farkli bir ad yanlislikla eslesmemeli
        assert yeni.katman_tipi("BA_WALL_XY") is None

    def test_kesfet_cli_config_yazar(self, yabanci_plan, tmp_path):
        from hakedis.cli import main

        hedef = tmp_path / "ofis.yml"
        assert main(["kesfet", str(yabanci_plan), "--cikti", str(hedef)]) == 0
        assert hedef.exists()
        # Yazilan config gercekten kullanilabilir olmali
        s, _ = plandan_metraj(str(yabanci_plan), ayarlari_yukle(hedef))
        assert len(s.tipe_gore(ElemanTipi.KOLON)) == 6
        assert len(s.tipe_gore(ElemanTipi.PERDE)) == 1


class TestHataYonetimi:
    def test_olmayan_dosya(self):
        with pytest.raises(FileNotFoundError):
            plandan_metraj("yok_boyle_bir_dosya.dxf", ayarlari_yukle())

    def test_desteklenmeyen_uzanti(self, tmp_path):
        p = tmp_path / "plan.docx"
        p.write_text("x")
        with pytest.raises(ValueError, match="Desteklenmeyen"):
            plandan_metraj(str(p), ayarlari_yukle())

    def test_bos_dxf_uyari_verir(self, tmp_path):
        import ezdxf

        doc = ezdxf.new("R2013")
        yol = tmp_path / "bos.dxf"
        doc.saveas(str(yol))
        sonuc, _ = plandan_metraj(str(yol), ayarlari_yukle())
        assert sonuc.elemanlar == []
        assert any("eleman tespit edilemedi" in u for u in sonuc.uyarilar)
